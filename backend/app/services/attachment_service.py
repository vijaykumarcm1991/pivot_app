"""
Attachment service — Phase 6.

Generates the .xlsx file that is attached to outgoing emails. The
service reuses the existing drilldown pipeline (`build_drilldown`
and `build_drilldown_multi` in `pivot_service`) so the rules for
"which raw records match a pivot row" are identical to the
drilldown modal in the UI. No duplicate logic.

Output:
  - A SheetJS-free, openpyxl-based .xlsx file with the matching
    raw records, one row per line, using the original column
    names and original values.
  - A meaningful filename like
        Pivot_Drilldown_2026-06-30_14-30.xlsx
  - The file is also persisted to REPORTS_DIR so the email-history
    page can offer a re-download.

The service is intentionally stateless — it does not know about
SMTP, history or recipients; those are wired together by
`email_service.compose_and_send`.
"""
from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.config.settings import REPORTS_DIR
from app.services.pivot_service import build_drilldown_multi


# Hard cap for the attachment. The same cap as the drilldown API
# (5 000) — keeps the file under ~1 MB even for wide sheets.
ATTACHMENT_LIMIT = 5000


class AttachmentError(ValueError):
    """Raised when the attachment cannot be generated."""


def _safe_filename_component(s: str, fallback: str = "dataset") -> str:
    """Strip filesystem-unfriendly characters from a string so it
    can be used in a filename."""
    if not s:
        return fallback
    keep = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_", "."):
            keep.append(ch)
        else:
            keep.append("_")
    out = "".join(keep).strip("_")
    return out or fallback


def default_filename(dataset_name: str = "", sheet_name: str = "") -> str:
    """Build a filename like `Pivot_Drilldown_2026-06-30_14-30.xlsx`."""
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    ds = _safe_filename_component(dataset_name, "dataset")
    # We intentionally do NOT include the sheet name in the
    # filename — the spec example only has the timestamp.
    return f"Pivot_Drilldown_{stamp}.xlsx"


def build_attachment(
    db: Session,
    dataset_id: int,
    sheet_name: str,
    rows: List[str],
    columns: List[str],
    values: List[Any],
    filters: Dict[str, Any],
    date_grouping: Dict[str, str],
    sorting: Dict[str, str],
    totals: Any,
    layout: str,
    selections: List[Dict[str, Any]],
    dataset_name: str = "",
    sheet_label: str = "",
) -> Tuple[bytes, str, int, int]:
    """
    Build the .xlsx attachment bytes + filename for the given
    pivot payload + list of selections.

    Returns a tuple of:
        (file_bytes, filename, matched_rows, returned_rows)
    """
    if not selections:
        raise AttachmentError("No pivot rows selected — nothing to attach.")

    drilldown = build_drilldown_multi(
        db=db,
        dataset_id=dataset_id,
        sheet_name=sheet_name,
        rows=rows,
        columns=columns,
        values=values,
        filters=filters,
        date_grouping=date_grouping,
        sorting=sorting,
        totals=totals,
        layout=layout,
        selections=selections,
        limit=ATTACHMENT_LIMIT,
    )

    rows_out: List[Dict[str, Any]] = drilldown.get("rows") or []
    columns_out: List[str] = list(drilldown.get("columns") or [])
    matched_rows = int(drilldown.get("metadata", {}).get("matched_rows") or 0)
    returned_rows = len(rows_out)

    if not columns_out and rows_out:
        # Defensive: if the engine returned rows but no column list,
        # derive it from the union of keys in the rows.
        seen = []
        for r in rows_out:
            for k in r.keys():
                if k not in seen:
                    seen.append(k)
        columns_out = seen

    file_bytes = _rows_to_xlsx_bytes(columns_out, rows_out)
    filename = default_filename(dataset_name=dataset_name, sheet_name=sheet_label)
    return file_bytes, filename, matched_rows, returned_rows


def save_attachment(
    file_bytes: bytes,
    filename: str,
    subdir: str = "email_attachments",
) -> str:
    """Persist the attachment to REPORTS_DIR/<subdir>/<filename> and
    return the on-disk path (relative to REPORTS_DIR). Used by the
    history page to offer a re-download link."""
    target_dir = os.path.join(REPORTS_DIR, subdir)
    os.makedirs(target_dir, exist_ok=True)
    on_disk = os.path.join(target_dir, filename)
    with open(on_disk, "wb") as f:
        f.write(file_bytes)
    return os.path.relpath(on_disk, REPORTS_DIR)


def attachment_disk_path(rel_path: Optional[str]) -> Optional[str]:
    """Resolve a stored relative path to an absolute path on disk,
    or None if the file no longer exists."""
    if not rel_path:
        return None
    # Belt-and-braces: reject anything that escapes REPORTS_DIR.
    abs_path = os.path.abspath(os.path.join(REPORTS_DIR, rel_path))
    if not abs_path.startswith(os.path.abspath(REPORTS_DIR) + os.sep):
        return None
    if not os.path.exists(abs_path):
        return None
    return abs_path


# ── Internals ────────────────────────────────────────────────────────────

def _rows_to_xlsx_bytes(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """Render a list of records to an in-memory .xlsx file using
    openpyxl. We deliberately avoid pulling in pandas here — the
    data is already in plain Python dicts and openpyxl gives us
    full control over the cell values and column ordering.

    Column order is honoured exactly as the engine returned it
    (which is also the order the user sees in the drilldown
    modal)."""

    wb = Workbook()
    ws = wb.active
    if ws is None:
        # openpyxl always gives us an active sheet on a fresh Workbook,
        # but be defensive in case that ever changes.
        ws = wb.create_sheet("Drilldown")
    ws.title = "Drilldown"

    # Header
    ws.append(list(columns))

    # Data
    for row in rows:
        ws.append([_xlsx_cell(row.get(c)) for c in columns])

    # Reasonable column widths: 12 chars default, capped at 40.
    for idx in range(1, len(columns) + 1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        header = str(columns[idx - 1]) if columns else ""
        width = max(12, min(40, len(header) + 2))
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_cell(value: Any) -> Any:
    """Convert a JSON-friendly value into something openpyxl can
    write to a cell. Returns None for `None`, the raw value for
    primitives, and the string form for unsupported types."""

    if value is None:
        return None
    if isinstance(value, bool):
        return bool(value)
    if isinstance(value, (int, float, str)):
        return value
    # datetimes / Decimals / other — openpyxl handles ISO strings for dates.
    try:
        return str(value)
    except Exception:
        return None
